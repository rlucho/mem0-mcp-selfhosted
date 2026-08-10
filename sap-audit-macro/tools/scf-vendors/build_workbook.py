"""Builds SCF_Vendor_Finder.xlsx -- the input side of the standalone tool.

The macro lives in modScfVendors.bas and is imported separately, because a
.xlsx carries no macros and this has to be buildable on a machine without
Excel.
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FONT = "Aptos Narrow"
HEAD = PatternFill("solid", fgColor="1F3864")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
GREY = PatternFill("solid", fgColor="F2F2F2")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)


def head(cell, text):
    cell.value = text
    cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    cell.fill = HEAD
    cell.alignment = Alignment(vertical="center")
    cell.border = THIN


SETTINGS = [
    ("Expected SAP system ID (SID)", "PP2",
     "The macro refuses to run if the attached session is a different system. Clear it to skip the check."),
    ("Company code", "GBHP",
     "Used for FB03 and for FBL1N. One company code per run."),
    ("Output folder", r"C:\Users\eslucres\Documents\SCF vendors",
     "Where the two exports per AB document are written. Created if missing."),
    ("Payment usage menu text", "Payment usage",
     "What Environment > Payment Usage is CALLED on your system. Found by this name first, "
     "because the recorded menu position is the G/L account master on the wrong screen -- "
     "which is how a run ends up in FS03."),
    ("Reference column", "Reference",
     "Heading of the column in the payment usage export that carries the ORIGINAL invoice "
     "number. This is the whole trick: the KA documents sit on the bank's account, and their "
     "Reference is the supplier invoice."),
    ("Invoice posting date from", "01012025",
     "FBL1N posting-date range for the supplier invoices. KEEP IT WIDE. The invoices are older "
     "than the payment -- in the batch this was built from the payment is September and the "
     "invoices run May to July. Too narrow and FBL1N finds nothing, which looks identical to "
     "'these documents do not exist'."),
    ("Invoice posting date to", "31122026", "The other end of that range."),
    ("Supplier column", "Supplier", "Heading of the vendor-account column in the FBL1N export."),
    ("Supplier name column", "Name 1", "Heading of the vendor-name column in that export."),
    ("Amount column", "Amount in local currency", "Heading of the amount column in that export."),
]

MAP = [
    ("FB03.DocNumber", "wnd[0]/usr/txtRF05L-BELNR"),
    ("FB03.CompanyCode", "wnd[0]/usr/ctxtRF05L-BUKRS"),
    ("FB03.FiscalYear", "wnd[0]/usr/txtRF05L-GJAHR"),
    ("PaymentUsage.Menu", "wnd[0]/mbar/menu[5]/menu[3]"),
    ("Export.ListMenu", "wnd[0]/mbar/menu[0]/menu[3]/menu[1]"),
    ("Save.Path", "wnd[1]/usr/ctxtDY_PATH"),
    ("Save.FileName", "wnd[1]/usr/ctxtDY_FILENAME"),
    ("Save.GenerateButton", "wnd[1]/tbar[0]/btn[0]"),
    ("Fbl1n.CompanyCode", "wnd[0]/usr/ctxtKD_BUKRS-LOW"),
    ("Fbl1n.AllItemsRadio", "wnd[0]/usr/radX_AISEL"),
    ("Fbl1n.PostingDateFrom", "wnd[0]/usr/ctxtSO_BUDAT-LOW"),
    ("Fbl1n.PostingDateTo", "wnd[0]/usr/ctxtSO_BUDAT-HIGH"),
    ("Fbl1n.DynamicSelections", "wnd[0]/tbar[1]/btn[16]"),
    ("Fbl1n.DocNumberField",
     "wnd[0]/usr/ssub%_SUBSCREEN_%_SUB%_CONTAINER:SAPLSSEL:2001/"
     "ssubSUBSCREEN_CONTAINER2:SAPLSSEL:2000/ssubSUBSCREEN_CONTAINER:SAPLSSEL:1106/"
     "txt%%DYN011-LOW"),
    ("Fbl1n.DocNumberMultiSelect",
     "wnd[0]/usr/ssub%_SUBSCREEN_%_SUB%_CONTAINER:SAPLSSEL:2001/"
     "ssubSUBSCREEN_CONTAINER2:SAPLSSEL:2000/ssubSUBSCREEN_CONTAINER:SAPLSSEL:1106/"
     "btn%_%%DYN011_%_APP_%-VALU_PUSH"),
    ("Fbl1n.ExecuteButton", "wnd[0]/tbar[1]/btn[8]"),
    ("MultiSel.PasteFromClipboard", "wnd[1]/tbar[0]/btn[24]"),
    ("MultiSel.Confirm", "wnd[1]/tbar[0]/btn[8]"),
]

EXAMPLE = [("995501", "2026"), ("995521", "2026")]

book = Workbook()

# --- Input ---------------------------------------------------------------
s = book.active
s.title = "Input"
s.sheet_properties.tabColor = "375623"
for col, width in {"A": 3, "B": 30, "C": 34, "D": 96}.items():
    s.column_dimensions[col].width = width

s["B2"] = "Who was paid in an SCF batch"
s["B2"].font = Font(name=FONT, size=15, bold=True, color="1F3864")
s["B3"] = (
    "A confirming settlement pays ONE vendor -- the bank -- and hides the suppliers behind it. "
    "Give this the AB payment document(s) and it walks FB03 > Payment Usage > the KA documents' "
    "Reference > FBL1N, and lists the suppliers actually paid."
)
s["B3"].font = Font(name=FONT, size=10, italic=True, color="595959")
s["B4"] = "Read-only: FB03 and FBL1N are display transactions. The only writes are files on your own PC."
s["B4"].font = Font(name=FONT, size=9, italic=True, color="808080")

head(s["B6"], "Setting")
head(s["C6"], "Value")
head(s["D6"], "Notes")
for i, (label, value, note) in enumerate(SETTINGS):
    r = 7 + i
    s.cell(r, 2, label).font = Font(name=FONT, size=10, bold=True)
    c = s.cell(r, 3, value)
    c.font = Font(name=FONT, size=10)
    c.fill = YELLOW
    c.border = THIN
    n = s.cell(r, 4, note)
    n.font = Font(name=FONT, size=9, color="595959")
    n.alignment = Alignment(wrap_text=True, vertical="top")
    s.row_dimensions[r].height = 30

first = 7 + len(SETTINGS) + 2
s.cell(first - 1, 2, "AB payment documents -- one per row").font = Font(
    name=FONT, size=11, bold=True, color="1F3864")
head(s.cell(first, 2), "AB document")
head(s.cell(first, 3), "Fiscal year")
head(s.cell(first, 4), "Result")
for i, (doc, year) in enumerate(EXAMPLE):
    r = first + 1 + i
    for col, value in ((2, doc), (3, year)):
        c = s.cell(r, col, value)
        c.fill = YELLOW
        c.border = THIN
        c.font = Font(name=FONT, size=10)
    c = s.cell(r, 4)
    c.fill = GREY
    c.border = THIN
for r in range(first + 1 + len(EXAMPLE), first + 40):
    for col in (2, 3):
        s.cell(r, col).fill = YELLOW
        s.cell(r, col).border = THIN
    s.cell(r, 4).fill = GREY
    s.cell(r, 4).border = THIN

s.cell(first + 41, 2, "The two example rows are the batch this was built from -- overwrite them.")
s.cell(first + 41, 2).font = Font(name=FONT, size=9, italic=True, color="808080")

# --- Vendors -------------------------------------------------------------
v = book.create_sheet("Vendors")
v.sheet_properties.tabColor = "C00000"
v["A2"] = "Suppliers actually paid"
v["A2"].font = Font(name=FONT, size=14, bold=True, color="1F3864")
v["A3"] = "Written by the macro. One row per supplier invoice found behind the AB document."
v["A3"].font = Font(name=FONT, size=9, italic=True, color="808080")
for i, (title, width) in enumerate([
    ("AB document", 14), ("Supplier", 12), ("Name", 34), ("Document", 14),
    ("Type", 7), ("Reference", 14), ("Document date", 15), ("Amount", 16),
]):
    head(v.cell(4, i + 1), title)
    v.column_dimensions[get_column_letter(i + 1)].width = width
v.freeze_panes = "A5"

# --- Log -----------------------------------------------------------------
g = book.create_sheet("Log")
g.sheet_properties.tabColor = "808080"
g["A2"] = "What happened"
g["A2"].font = Font(name=FONT, size=14, bold=True, color="1F3864")
for i, (title, width) in enumerate([
    ("Timestamp", 20), ("AB document", 14), ("Step", 18), ("Detail", 130)]):
    head(g.cell(4, i + 1), title)
    g.column_dimensions[get_column_letter(i + 1)].width = width
g.freeze_panes = "A5"

# --- Screen Map ----------------------------------------------------------
m = book.create_sheet("Screen Map")
m.sheet_properties.tabColor = "BF8F00"
m["A1"] = "Key"
m["B1"] = "SAP element ID"
for cell in (m["A1"], m["B1"]):
    cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    cell.fill = HEAD
m.column_dimensions["A"].width = 30
m.column_dimensions["B"].width = 104
for i, (key, value) in enumerate(MAP):
    m.cell(i + 2, 1, key).font = Font(name=FONT, size=10)
    c = m.cell(i + 2, 2, value)
    c.font = Font(name=FONT, size=9)
    c.fill = YELLOW
m.cell(len(MAP) + 3, 1, "Blank or missing rows fall back to these same values, which are what a "
                        "live PP2 session answered to. Correct one here if a screen has moved.")
m.cell(len(MAP) + 3, 1).font = Font(name=FONT, size=9, italic=True, color="808080")

book.save("tools/scf-vendors/SCF_Vendor_Finder.xlsx")
print("wrote tools/scf-vendors/SCF_Vendor_Finder.xlsx")
