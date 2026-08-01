#!/usr/bin/env python3
"""Build FEBAN_Audit_Control.xlsx -- the workbook the VBA macro is driven from.

The macro holds no SAP screen-element IDs of its own. Every ID it touches is read
from the 'Screen Map' sheet, so adapting the macro to this SAP release is a
paste-the-recording exercise rather than a code change.

Sheets produced:
    Control      settings the operator fills in (system, company code, paths, mode)
    Screen Map   SAP element IDs harvested from an Alt+F12 script recording
    Samples      the 56 audit samples, normalised, plus columns the macro writes
    Log          empty audit trail; the macro appends one row per action
    Data Issues  data-quality flags raised while reading the auditor's workbook

Usage:
    python3 build_control_workbook.py [-i samples.csv] [-o FEBAN_Audit_Control.xlsx]
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FONT = "Arial"

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name=FONT, size=10, bold=True, color="FFFFFF")
SECTION_FONT = Font(name=FONT, size=11, bold=True, color="1F3864")
BODY_FONT = Font(name=FONT, size=10)
INPUT_FONT = Font(name=FONT, size=10, color="0000FF")
NOTE_FONT = Font(name=FONT, size=9, italic=True, color="595959")

INPUT_FILL = PatternFill("solid", fgColor="FFFF00")     # operator fills these in
MACRO_FILL = PatternFill("solid", fgColor="EDEDED")     # macro writes these
EXAMPLE_FILL = PatternFill("solid", fgColor="FFF2CC")   # illustrative example row

THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DATE_FMT = "DD/MM/YYYY"
AMOUNT_FMT = "#,##0.00"


def style_header(sheet, row: int, last_col: int) -> None:
    for col in range(1, last_col + 1):
        cell = sheet.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BOX
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = sheet.cell(row=row + 1, column=1)


def set_widths(sheet, widths: dict[str, int]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


# --------------------------------------------------------------------------- #
# Control
# --------------------------------------------------------------------------- #

CONTROL_SETTINGS = [
    ("Expected SAP system ID (SID)", "PP2",
     "The macro aborts if the attached session reports a different SID. Stops an "
     "audit extract being run against the wrong system."),
    ("Expected SAP client", "",
     "Optional. Leave blank to skip the client check."),
    ("Company code", "GBKM",
     "Per the audit request."),
    ("Transaction for statement search", "FEBAN",
     "FEBAN is a POST-PROCESSING transaction, not a display-only one. See README, "
     "'Why FEBAN is not read-only'. Set to FF.6 for a display-only statement view."),
    ("House bank", "",
     "Optional FEBAN filter. Leave blank for all house banks."),
    ("Account ID", "",
     "Optional FEBAN filter. Leave blank for all accounts."),
    ("Download root folder", r"C:\Users\eslucres\Documents\Audit GBKM",
     "Taken from recordings/Audit.vbs. Created if missing. One subfolder per month tab, "
     "e.g. ...\\Audit GBKM\\Sep 25."),
    ("Confirming party name", "SANTANDER SCF",
     "If the largest ZP payment of the batch is to this party, it is a confirming "
     "(supply-chain-finance) payment and needs the extra hop to reach the supplier "
     "invoices. Matched on letters and digits, so 'SCF Santander' also matches, but "
     "'SANTANDER UK PLC' does not."),
    ("Run mode", "DRY RUN",
     "Both modes run the WHOLE chain and write the same files -- the exports are "
     "read-only on the SAP side, and the chain reads them back, so it cannot run "
     "without them. The only difference is where they land: DRY RUN puts everything "
     "under a '_dry run' subfolder of the download root, so a rehearsal is never "
     "mistaken for the evidence pack. Do one clean DRY RUN, read the Log, then switch."),
    ("Amount match tolerance", 0.01,
     "Absolute currency tolerance when matching a sample amount to a statement line."),
    ("Stop on first error", "YES",
     "YES halts the run on the first unrecognised screen. NO logs and continues to "
     "the next sample."),
    ("Max seconds to wait per screen", 60,
     "Guards against an indefinite hang when SAP is slow."),
    ("Max rows for a settlement", 8,
     "A clearing document with no vendor payments and no more than this many "
     "bookkeeping rows is a treasury, tax or FX settlement -- there is no invoice "
     "behind it. Above it, the run assumes the document-type column was misread. "
     "A real payment run here has held between 34 and 656 payments, never a handful."),
    ("SAP date format", "DDMMYYYY",
     "How dates are typed into SAP. DDMMYYYY = 01092025, which is what "
     "recordings/Audit.vbs used and therefore what is known to work here. Also "
     "supported: DMY = 31.12.2025, DMY/ = 31/12/2025, MDY = 12/31/2025, "
     "YMD = 2025-12-31. Get this wrong and FEBAN silently searches the wrong period."),
    ("Payment document type", "ZP",
     "Only documents of these types are taken from the Payment Usage list and fed to "
     "FBL1N. ZP is the SAP standard for a payment document. Takes a comma-separated "
     "list -- 'ZP, ZV, KZ' -- for batches paid outside the normal payment run. When "
     "nothing matches, the Log names the types the file actually held."),
    ("Clearing line posting key", "40",
     "Step 4 opens the first line-item row with this posting key that also carries a "
     "clearing document. Blank means 'any row with a clearing document'."),
    ("Payment usage document column", "",
     "Caption of the document-number column in the exported Payment Usage list, if the "
     "macro cannot find it. Open the first exported file and copy the heading exactly."),
    ("Payment usage type column", "",
     "Caption of the document-type column in that same file. Without it every document "
     "is taken rather than only the ZP ones, and the Log says so."),
    ("ZP list amount column", "",
     "Caption of the amount column in the exported FBL1N payment list. Leave blank -- the "
     "macro tries the caption, then the SAP technical name, then works it out from what "
     "the column contains, which is language-independent. Only fill this in if the Log "
     "says it picked the wrong one."),
    ("ZP list vendor column", "",
     "Caption of the vendor-name column in that same file. Same rules."),
    ("ZP list document column", "",
     "Caption of the document-number column in that same file. Same rules."),
    ("Invoice attachment title contains", "Invoice",
     "An invoice document carries more than one attachment -- on this system the "
     "workflow notes sit above the document itself -- so the row whose text contains "
     "this word is the one downloaded. The titles come from the archiving system rather "
     "than from SAP, so they do not follow the SAP logon language. Nothing matching "
     "takes the last row and says so in the Log, naming every attachment it saw."),
    ("Invoice document type", "KR, RN",
     "CROSS-CHECK ONLY -- this no longer decides anything. The invoice is picked by SIGN: "
     "in a vendor line-item list the payment is a debit and the invoice it settles is a "
     "credit, so the invoice is the negative row and the biggest invoice is the most "
     "negative one. That holds whatever the type is called in this company code, which "
     "matters because it is RN here and KR on the SAP standard. If the row picked is not "
     "one of the types listed, the Log says so and takes it anyway. Blank to switch the "
     "cross-check off."),
    ("Invoice list amount column", "",
     "Santander SCF route only. Caption of the amount column in the exported list of "
     "invoices behind an SCF payment. Blank falls back to a built-in list of captions."),
    ("Invoice list supplier column", "",
     "Caption of the supplier column in that same file."),
    ("Invoice list document column", "",
     "Caption of the invoice-number column in that same file."),
    ("Extra popup titles to allow", "",
     "Pipe-separated. The run stops on any modal popup it does not recognise rather than "
     "pressing Enter through it. The save dialog and the attachment list are identified by "
     "their contents, so language does not matter for those -- add a title here only if "
     "your system shows a popup the run does not expect."),
    ("Operator name", "",
     "Written into the log for the audit trail."),
]


def build_control(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Control")
    sheet.sheet_properties.tabColor = "1F3864"
    set_widths(sheet, {"A": 3, "B": 38, "C": 30, "D": 78})

    sheet["B2"] = "FEBAN audit extract - control sheet"
    sheet["B2"].font = Font(name=FONT, size=14, bold=True, color="1F3864")
    sheet["B3"] = (
        "Fill in the yellow cells, paste your recorded element IDs into 'Screen Map', "
        "then run modMain.RunExtract from the VBA editor."
    )
    sheet["B3"].font = NOTE_FONT

    sheet["B5"] = "Legend"
    sheet["B5"].font = SECTION_FONT
    legend = [
        ("Yellow fill", "You fill this in. The macro reads it and never overwrites it."),
        ("Grey fill", "The macro writes here. Do not edit -- your edits are overwritten each run."),
        ("Blue text", "A hardcoded input value."),
        ("Cream fill", "An illustrative example, not live data. Delete or overwrite it."),
    ]
    for offset, (label, meaning) in enumerate(legend):
        row = 6 + offset
        sheet[f"B{row}"] = label
        sheet[f"B{row}"].font = BODY_FONT
        sheet[f"C{row}"] = meaning
        sheet[f"C{row}"].font = NOTE_FONT
        sheet[f"C{row}"].alignment = Alignment(wrap_text=True)
    sheet["B6"].fill = INPUT_FILL
    sheet["B7"].fill = MACRO_FILL
    sheet["B8"].font = INPUT_FONT
    sheet["B9"].fill = EXAMPLE_FILL

    header_row = 12
    sheet[f"B{header_row}"] = "Setting"
    sheet[f"C{header_row}"] = "Value"
    sheet[f"D{header_row}"] = "Notes"
    for col in ("B", "C", "D"):
        cell = sheet[f"{col}{header_row}"]
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BOX
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # The macro locates settings by matching column B, so these labels are the
    # contract between this sheet and modConfig. Renaming one breaks the lookup.
    for offset, (label, value, note) in enumerate(CONTROL_SETTINGS):
        row = header_row + 1 + offset
        sheet[f"B{row}"] = label
        sheet[f"B{row}"].font = BODY_FONT
        sheet[f"B{row}"].border = BOX

        cell = sheet[f"C{row}"]
        cell.value = value
        cell.font = INPUT_FONT
        cell.fill = INPUT_FILL
        cell.border = BOX
        if isinstance(value, float):
            cell.number_format = AMOUNT_FMT

        sheet[f"D{row}"] = note
        sheet[f"D{row}"].font = NOTE_FONT
        sheet[f"D{row}"].border = BOX
        sheet[f"D{row}"].alignment = Alignment(wrap_text=True, vertical="top")

    mode_row = header_row + 1 + [s[0] for s in CONTROL_SETTINGS].index("Run mode")
    mode_validation = DataValidation(
        type="list", formula1='"DRY RUN,EXTRACT"', allow_blank=False, showDropDown=False
    )
    sheet.add_data_validation(mode_validation)
    mode_validation.add(sheet.cell(row=mode_row, column=3))
    sheet.cell(row=mode_row, column=3).comment = Comment(
        "DRY RUN is the safe default. It navigates and logs but exports nothing.",
        "build_control_workbook.py",
    )

    for label in ("Stop on first error",):
        row = header_row + 1 + [s[0] for s in CONTROL_SETTINGS].index(label)
        validation = DataValidation(
            type="list", formula1='"YES,NO"', allow_blank=False, showDropDown=False
        )
        sheet.add_data_validation(validation)
        validation.add(sheet.cell(row=row, column=3))

    progress_row = header_row + len(CONTROL_SETTINGS) + 3
    sheet[f"B{progress_row}"] = "Progress (formulas over the Samples sheet)"
    sheet[f"B{progress_row}"].font = SECTION_FONT
    # One counter per status modMain can write, so the sheet reflects the run
    # rather than only its successes.
    progress = [
        ("Samples in the sheet", "=COUNTA(Samples!A5:A1000)"),
        ("Included in the next run", '=COUNTIF(Samples!P5:P1000,"Yes")+COUNTBLANK(Samples!P5:P60)'),
        ("Excluded", '=COUNTIF(Samples!P5:P1000,"No")'),
        ("DONE -- invoice downloaded", '=COUNTIF(Samples!J5:J1000,"DONE")'),
        ("BLOCKED_FBL1N -- steps 8-9 not mapped yet",
         '=COUNTIF(Samples!J5:J1000,"BLOCKED_FBL1N")'),
        ("BLOCKED_INVOICE -- reached the payment, no PDF",
         '=COUNTIF(Samples!J5:J1000,"BLOCKED_INVOICE")'),
        ("PARTIAL -- traced, no invoice file", '=COUNTIF(Samples!J5:J1000,"PARTIAL")'),
        # An FI document that clears nothing settles no supplier, so there is
        # no invoice to fetch. Its line items are the evidence instead. That
        # is a complete answer, not a failure, and it gets its own counter so
        # it stops being read as one.
        ("NO CLEARING -- internal transfer, line items exported",
         '=COUNTIF(Samples!J5:J1000,"NO CLEARING")'),
        # Settles against the bank statement, not a supplier -- an FX or
        # treasury movement. No invoice exists to fetch.
        ("NO VENDOR PAYMENTS -- treasury/FX settlement",
         '=COUNTIF(Samples!J5:J1000,"NO VENDOR PAYMENTS")'),
        ("NOT FOUND -- no matching statement line",
         '=COUNTIF(Samples!J5:J1000,"NOT FOUND")'),
        ("AMBIGUOUS -- more than one line matched",
         '=COUNTIF(Samples!J5:J1000,"AMBIGUOUS")'),
        ("ERROR", '=COUNTIF(Samples!J5:J1000,"ERROR")'),
        ("Not started", "=COUNTBLANK(Samples!J5:J60)"),
        ("Files downloaded", "=SUM(Samples!N5:N1000)"),
        ("Samples with no named beneficiary in the request",
         '=COUNTIF(Samples!I5:I1000,"*no_named_beneficiary*")'),
        ("Reached a confirming (Santander SCF) payment",
         '=COUNTIF(Samples!O5:O1000,"*confirming party*")'),
        ("Samples the request already names as SANTANDER SCF",
         '=COUNTIF(Samples!G5:G1000,"*SANTANDER*")'),
    ]
    for offset, (label, formula) in enumerate(progress):
        row = progress_row + 1 + offset
        sheet[f"B{row}"] = label
        sheet[f"B{row}"].font = BODY_FONT
        sheet[f"C{row}"] = formula
        sheet[f"C{row}"].font = BODY_FONT
        sheet[f"C{row}"].fill = MACRO_FILL
        sheet[f"C{row}"].border = BOX

    assumptions_row = progress_row + len(progress) + 3
    sheet[f"B{assumptions_row}"] = "Assumptions baked into this workbook"
    sheet[f"B{assumptions_row}"].font = SECTION_FONT
    assumptions = [
        "Company code GBKM and the FEBAN transaction are taken from the requester's "
        "instruction, not from the auditor's workbook, which names neither.",
        "The FEBAN statement-date range for each sample is the first to the last "
        "calendar day of the month containing that sample's payment date "
        "(column C and D formulas on 'Samples').",
        "Sample amounts are stored as positive numbers. The auditor's workbook shows "
        "them unsigned even where the bank statement shows a debit.",
        "Source of the sample list: 'Paper Samples' sheet, rows 5-60, of "
        "Samples_Paper_SURL260716_152455.962.xlsx as supplied by the auditor.",
    ]
    for offset, text in enumerate(assumptions):
        row = assumptions_row + 1 + offset
        sheet[f"B{row}"] = f"{offset + 1}."
        sheet[f"B{row}"].font = BODY_FONT
        sheet[f"C{row}"] = text
        sheet[f"C{row}"].font = NOTE_FONT
        sheet[f"C{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        sheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        sheet.row_dimensions[row].height = 28


# --------------------------------------------------------------------------- #
# Screen Map
# --------------------------------------------------------------------------- #

# key, what it is, ID, required?
#
# The IDs below are the ones actually captured in recordings/Audit.vbs on this
# system, not generic examples -- 'recorded' in the Source column marks those.
# 'VERIFY' marks a row the recording did not cover, where the value is a
# standard-but-unconfirmed guess that has to be checked before an EXTRACT run.
SCREEN_MAP_ROWS = [
    ("--- Step 1: FEBAN selection (a MODAL popup on this system, wnd[1]) ---", "", "", ""),
    ("FEBAN.SelectionWindow", "Window the selection fields sit in. wnd[1] here, because "
     "FEBAN opens its selection as a popup rather than a full screen", "wnd[1]", "Yes"),
    ("FEBAN.CompanyCode", "Company code", "wnd[1]/usr/ctxtSL_BUKRS-LOW", "Yes"),
    ("FEBAN.StatementDateFrom", "Statement date, low", "wnd[1]/usr/ctxtSL_AZDAT-LOW", "Yes"),
    ("FEBAN.StatementDateTo", "Statement date, high", "wnd[1]/usr/ctxtSL_AZDAT-HIGH", "Yes"),
    ("FEBAN.ExecuteButton", "Execute, on the popup's own toolbar",
     "wnd[1]/tbar[0]/btn[8]", "Yes"),
    ("FEBAN.PostExecuteButton", "Toolbar button pressed straight after Execute in all four "
     "recordings. Clear it to skip", "wnd[0]/tbar[1]/btn[14]", "No"),
    ("FEBAN.HouseBank", "House bank filter. Not recorded; worth setting if the EUR "
     "account is out of scope. VERIFY", "wnd[1]/usr/ctxtSL_HBKID-LOW", "No"),
    ("FEBAN.AccountId", "Account ID filter. VERIFY", "wnd[1]/usr/ctxtSL_HKTID-LOW", "No"),
    ("--- Step 2: FEBAN result list ---", "", "", ""),
    ("FEBAN.ResultGrid", "ALV grid of statement items", "wnd[0]/shellcont/shell", "Yes"),
    ("FEBAN.Col.Amount", "Amount column", "KWBTR", "Yes"),
    ("FEBAN.Col.ValueDate", "Statement date column. Audit2.vbs double-clicked this by "
     "name, so it is a captured value", "AZDAT", "Yes"),
    ("FEBAN.Col.Status", "Update-status column. The grid dump on PP2 shows no ESTAT -- "
     "VB1OK 'Update 1 OK' and VB2OK '2nd update OK' are what this release carries",
     "VB1OK", "No"),
    ("FEBAN.Col.DocNumber", "FI document column. VERIFY", "BELNR", "No"),
    ("FEBAN.Col.Reference", "Note-to-payee, logged beside each match as context. The grid "
     "dump shows no SGTXT; VWEZW is 'Payment Notes' and VGREF is 'Bank reference'",
     "VWEZW", "No"),
    ("--- ALV export (confirmed in Audit2.vbs) ---", "", "", ""),
    ("Export.AlvToolbarButton", "Grid toolbar export function code",
     "&MB_EXPORT", "No"),
    ("Export.AlvMenuItem", "Context-menu entry. &XXL, not &PC", "&XXL", "No"),
    ("Export.AlvFormatRadio", "Radio for the format, if that popup appears. VERIFY",
     "", "No"),
    ("--- Step 3: statement item detail ---", "", "", ""),
    ("Feban.Detail.DocNumber", "Posting Area 1 Doc. number field. F2 from here opens the "
     "FI document",
     "wnd[0]/usr/subSUB_MAIN:SAPLNEW_FEBA:4000/subSUB_APPLICATION:SAPLNEW_FEBA:2200/"
     "subAREA1:SAPLNEW_FEBA:2201/txtD2201_BELNR", "Yes"),
    ("--- Step 4: FI document line items ---", "", "", ""),
    ("Doc.BsegGrid", "Line-item grid on the document overview",
     "wnd[0]/usr/cntlCTRL_CONTAINERBSEG/shellcont/shell", "Yes"),
    ("Doc.Col.ClearingDoc", "Clearing-document column. Audit3 and Audit5 both "
     "double-clicked AUGBL; Audit.vbs used DMBTR and Audit2 PRCTR, which were just where "
     "the cursor sat", "AUGBL", "Yes"),
    ("Doc.Col.PostingKey", "Posting-key column, so the macro can pick the key-40 line "
     "rather than whichever row the cursor lands on. VERIFY", "BSCHL", "No"),
    ("--- Step 5: line-item detail ---", "", "", ""),
    ("Doc.ClearingDocField", "Clearing-document field. F2 from here opens the clearing "
     "document", "wnd[0]/usr/txtBSEG-AUGBL", "Yes"),
    ("Doc.OverviewButton", "Document-overview button. F2 on the FBL1N list lands on the "
     "LINE ITEM, and Environment > Payment usage lives on the document -- this is the hop "
     "between them, pressed in N1.vbs. Only used when the usage menu is missing, so "
     "clearing it just skips the attempt", "wnd[0]/tbar[1]/btn[9]", "No"),
    ("--- Step 6: Environment > Payment Usage ---", "", "", ""),
    ("PaymentUsage.Menu", "Menu path to the batch's payment documents. Confirmed in all "
     "four recordings", "wnd[0]/mbar/menu[5]/menu[3]", "Yes"),
    ("PaymentUsage.ListAnchor", "An element that exists on that list, used only to confirm "
     "the menu landed. Clear it to skip the check", "wnd[0]/usr/chk[1,6]", "No"),
    ("--- Step 7: classic list export (List > Save/Send > File) ---", "", "", ""),
    ("Export.ListMenu", "Menu path that opens the save-list dialog",
     "wnd[0]/mbar/menu[0]/menu[3]/menu[1]", "Yes"),
    ("Export.FormatOkButton", "Confirm the file-format popup", "wnd[1]/tbar[0]/btn[0]", "Yes"),
    ("Save.Path", "Directory field", "wnd[1]/usr/ctxtDY_PATH", "Yes"),
    ("Save.FileName", "File name field. Confirmed in Audit2.vbs and Audit5.vbs",
     "wnd[1]/usr/ctxtDY_FILENAME", "Yes"),
    ("Save.Encoding", "Encoding field, where present. VERIFY",
     "wnd[1]/usr/ctxtDY_FILE_ENCODING", "No"),
    ("Save.GenerateButton", "Confirm the save. btn[0], not btn[11]",
     "wnd[1]/tbar[0]/btn[0]", "Yes"),
    ("--- Steps 8-9: FBL1N, the ZP payments of the batch ---", "", "", ""),
    ("Fbl1n.AllItemsRadio", "'All items' radio", "wnd[0]/usr/radX_AISEL", "No"),
    ("Fbl1n.CompanyCode", "Company code. KD_BUKRS, not the DD_BUKRS that was predicted",
     "wnd[0]/usr/ctxtKD_BUKRS-LOW", "No"),
    ("Fbl1n.PostingDateFrom", "Posting date, low", "wnd[0]/usr/ctxtSO_BUDAT-LOW", "No"),
    ("Fbl1n.PostingDateTo", "Posting date, high", "wnd[0]/usr/ctxtSO_BUDAT-HIGH", "No"),
    ("Fbl1n.DynamicSelections", "Opens the dynamic-selections block. The document number "
     "is NOT on the main selection screen -- it lives in dynamic selections, so this has "
     "to be pressed before the field below exists", "wnd[0]/tbar[1]/btn[16]", "No"),
    ("Fbl1n.DocNumberField", "The document-number field itself, inside dynamic "
     "selections. Used when the batch holds a single document, so the multiple-selection "
     "dialog is not needed",
     "wnd[0]/usr/ssub%_SUBSCREEN_%_SUB%_CONTAINER:SAPLSSEL:2001/"
     "ssubSUBSCREEN_CONTAINER2:SAPLSSEL:2000/ssubSUBSCREEN_CONTAINER:SAPLSSEL:1106/"
     "txt%%DYN011-LOW", "No"),
    ("Fbl1n.DocNumberMultiSelect", "Multiple-selection arrow for document number, inside "
     "the dynamic-selections subscreen",
     "wnd[0]/usr/ssub%_SUBSCREEN_%_SUB%_CONTAINER:SAPLSSEL:2001/"
     "ssubSUBSCREEN_CONTAINER2:SAPLSSEL:2000/ssubSUBSCREEN_CONTAINER:SAPLSSEL:1106/"
     "btn%_%%DYN011_%_APP_%-VALU_PUSH", "No"),
    ("MultiSel.PasteFromClipboard", "'Upload from clipboard' on the multiple-selection "
     "dialog", "wnd[1]/tbar[0]/btn[24]", "No"),
    ("MultiSel.Confirm", "Copy on that dialog", "wnd[1]/tbar[0]/btn[8]", "No"),
    ("Fbl1n.ExecuteButton", "Execute", "wnd[0]/tbar[1]/btn[8]", "No"),
    ("Fbl1n.ResultAnchor", "Confirms the FBL1N result came up. NOTE: FBL1N renders as a "
     "CLASSIC LIST here (lbl[x,y]), not an ALV grid, so there is no grid to read -- the "
     "list is exported and read back instead. VERIFY", "wnd[0]/usr/lbl[164,8]", "No"),
    ("--- FB03: fallback only, not the normal path ---", "", "", ""),
    ("FB03.DocNumber", "The largest payment is normally opened from the FBL1N list "
     "itself, by finding the label whose text is its document number -- the same F2 "
     "drill the recordings use, aimed by content instead of by screen position. FB03 is "
     "only used if that number is not on screen, e.g. a long list that has scrolled. "
     "Leave these blank to disable the fallback entirely. VERIFY",
     "wnd[0]/usr/txtRF05L-BELNR", "No"),
    ("FB03.CompanyCode", "Company code field on that screen. VERIFY",
     "wnd[0]/usr/ctxtRF05L-BUKRS", "No"),
    ("FB03.FiscalYear", "Fiscal year field on that screen. VERIFY",
     "wnd[0]/usr/txtRF05L-GJAHR", "No"),
    ("--- Step 10: the largest payment's invoices, and the PDF ---", "", "", ""),
    ("Payment.UsageMenu", "Environment > Payment usage from inside the payment document. "
     "menu[4] here, not the menu[5] used from the clearing document -- menu indices differ "
     "per screen. VERIFY", "wnd[0]/mbar/menu[4]/menu[3]", "No"),
    ("Invoice.GosToolbox", "Services-for-object toolbox. The container index is NOT "
     "stable -- N1.vbs found it at shellcont[2] and B2.vbs at shellcont[1] on the same "
     "system, because it depends on what else the title bar carries. The macro tries "
     "this value first and then probes the neighbours, so it is a hint, not a hard ID",
     "wnd[0]/titl/shellcont[1]/shell", "No"),
    ("Invoice.AttachListGrid", "Attachment-list grid. CONTAINER_0100",
     "wnd[1]/usr/cntlCONTAINER_0100/shellcont/shell", "No"),
    ("Invoice.AttachListColumn", "Column used to select the attachment row",
     "BITM_DESCR", "No"),
    ("Invoice.ExportMenuItem", "Context-menu entry that saves the attachment. A CONTEXT "
     "MENU item, not the toolbar button that was predicted", "%ATTA_EXPORT", "No"),
    ("Invoice.SaveWindow", "Window the PDF save dialog opens in. wnd[2], because the "
     "attachment list is already wnd[1]", "wnd[2]", "No"),
    ("Invoice.CloseAttachList", "Closes the attachment list afterwards",
     "wnd[1]/tbar[0]/btn[0]", "No"),
]


def build_screen_map(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Screen Map")
    sheet.sheet_properties.tabColor = "C55A11"
    set_widths(sheet, {"A": 3, "B": 30, "C": 52, "D": 62, "E": 11, "F": 62})

    sheet["B2"] = "SAP element IDs"
    sheet["B2"].font = Font(name=FONT, size=14, bold=True, color="1F3864")
    sheet["B3"] = (
        "Column F is what the macro reads. Column D says how much to trust it. "
        "recorded = came straight out of a recording in recordings/, so it should be "
        "right for this system. VERIFY = standard SAP, but no recording touched it. "
        "PREDICTED = written from standard SAP for steps 8-10, which no recording "
        "reaches at all -- put SAP on the screen in question and run "
        "modProbe.ProbeCurrentScreen to find out which of these hold, then correct them "
        "here. BLOCKED = genuinely unknown, and the run stops short of that stage and "
        "says so."
    )
    sheet["B3"].font = NOTE_FONT
    sheet["B3"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet.merge_cells("B3:F3")
    sheet.row_dimensions[3].height = 56

    header_row = 5
    headers = ["Key", "What it is", "Source", "Required", "Element ID (the macro reads this)"]
    for offset, title in enumerate(headers):
        sheet.cell(row=header_row, column=2 + offset, value=title)
    style_header(sheet, header_row, 6)

    row = header_row + 1
    for key, description, element_id, required in SCREEN_MAP_ROWS:
        if key.startswith("---"):
            sheet.cell(row=row, column=2, value=key.strip("- ").strip())
            sheet.cell(row=row, column=2).font = SECTION_FONT
            sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            for col in range(2, 7):
                sheet.cell(row=row, column=col).fill = PatternFill("solid", fgColor="DEEAF6")
                sheet.cell(row=row, column=col).border = BOX
            row += 1
            continue

        sheet.cell(row=row, column=2, value=key).font = Font(name=FONT, size=10, bold=True)
        sheet.cell(row=row, column=3, value=description).font = NOTE_FONT
        sheet.cell(row=row, column=3).alignment = Alignment(wrap_text=True, vertical="top")

        # Provenance, so nobody mistakes a standard guess for a captured value.
        if "BLOCKED" in description:
            source, colour = "BLOCKED", "C00000"
        elif "PREDICTED" in description:
            # Written from standard SAP, never observed on this system. Must be
            # checked with modProbe.ProbeCurrentScreen before an EXTRACT run.
            source, colour = "PREDICTED", "C55A11"
        elif "VERIFY" in description or "NOT in the recording" in description:
            source, colour = "VERIFY", "BF8F00"
        elif element_id:
            source, colour = "recorded", "375623"
        else:
            source, colour = "optional", "595959"

        source_cell = sheet.cell(row=row, column=4, value=source)
        source_cell.font = Font(name=FONT, size=9, bold=True, color=colour)
        source_cell.alignment = Alignment(horizontal="center")

        required_cell = sheet.cell(row=row, column=5, value=required)
        required_cell.font = BODY_FONT
        required_cell.alignment = Alignment(horizontal="center")

        target = sheet.cell(row=row, column=6, value=element_id or None)
        target.fill = INPUT_FILL
        target.font = Font(name=FONT, size=8, color="0000FF")
        target.alignment = Alignment(wrap_text=True, vertical="top")

        for col in range(2, 7):
            sheet.cell(row=row, column=col).border = BOX
        if element_id and len(element_id) > 70:
            sheet.row_dimensions[row].height = 30
        row += 1

    note_row = row + 1
    for text in (
        "The macro reads column F only. A blank 'Required = Yes' row aborts the run with a "
        "message naming the missing key, rather than guessing at an ID.",
        "FEBAN's selection screen is a modal popup on this system (wnd[1]), not a full "
        "screen. That is why FEBAN.SelectionWindow exists -- the macro has to know which "
        "window to press Execute in.",
        "The recording entered dates as 8 digits with no separators (01092025). The Control "
        "sheet's 'SAP date format' is set to DDMMYYYY to match. Do not change it to DMY "
        "unless you have checked that this system accepts 01.09.2025 as well.",
    ):
        sheet.cell(row=note_row, column=2, value=text).font = NOTE_FONT
        sheet.cell(row=note_row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        sheet.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=6)
        sheet.row_dimensions[note_row].height = 30
        note_row += 1


# --------------------------------------------------------------------------- #
# Samples
# --------------------------------------------------------------------------- #

SAMPLE_HEADERS = [
    ("A", "#", 5),
    ("B", "Month tab", 11),
    ("C", "Stmt date from", 14),
    ("D", "Stmt date to", 14),
    ("E", "Payment date", 14),
    ("F", "Amount", 15),
    ("G", "Name of party", 26),
    ("H", "Payment reference", 27),
    ("I", "Data flags (from the auditor's file)", 44),
    ("J", "Status", 11),
    ("K", "Bank stmt / FEBAN item", 22),
    ("L", "FI document", 16),
    ("M", "Vendor invoice(s)", 24),
    ("N", "Files", 7),
    ("O", "Message", 52),
    ("P", "Include?", 10),
    # Added when the macro stopped being about one audit request. Appended
    # rather than inserted, so the Control sheet's formulas over A:P still
    # point at the same columns.
    ("Q", "Request", 30),
    ("R", "Company code", 13),
    ("S", "Auditor's comment", 46),
    # Set only when the row came from a SAP document list: the document to
    # enter the chain at, and which rung that is. Blank on every row that
    # came from an auditor's statement extract, which is the normal case.
    ("T", "Start document", 16),
    ("U", "Start at", 12),
    # A follow-up request pastes the SAP extract it worked from under each
    # sample, naming the ZP the auditor treated as the largest payment of the
    # batch. Captured so the run can be checked against their answer. Blank
    # whenever the file shows no working, which is most of them.
    ("V", "Auditor's ZP", 14),
]
MACRO_COLUMNS = ("J", "K", "L", "M", "N", "O")


def build_samples(workbook: Workbook, rows: list[dict]) -> None:
    sheet = workbook.create_sheet("Samples")
    sheet.sheet_properties.tabColor = "375623"
    set_widths(sheet, {col: width for col, _, width in SAMPLE_HEADERS})

    sheet["A2"] = "Audit samples"
    sheet["A2"].font = Font(name=FONT, size=14, bold=True, color="1F3864")
    sheet["A3"] = (
        "Use 'Import request' to add an auditor's sample file -- it finds the header "
        "wherever it is and asks which company code the request belongs to. Columns C and "
        "D are formulas, so correcting a payment date in column E re-derives the FEBAN "
        "statement-date range automatically."
    )
    sheet["A3"].font = NOTE_FONT
    sheet.merge_cells("A3:I3")

    header_row = 4
    for column, title, _ in SAMPLE_HEADERS:
        sheet[f"{column}{header_row}"] = title
    style_header(sheet, header_row, len(SAMPLE_HEADERS))

    for offset, record in enumerate(rows):
        row = header_row + 1 + offset

        sheet[f"A{row}"] = int(record["idx"])
        sheet[f"B{row}"] = record["month_tab"]

        # Derived, not copied, so the range always matches the payment date.
        sheet[f"C{row}"] = f"=IF($E{row}=\"\",\"\",DATE(YEAR($E{row}),MONTH($E{row}),1))"
        sheet[f"D{row}"] = f"=IF($E{row}=\"\",\"\",EOMONTH($E{row},0))"

        payment_date = record["payment_date"]
        sheet[f"E{row}"] = dt.date.fromisoformat(payment_date) if payment_date else None
        sheet[f"F{row}"] = float(record["amount"]) if record["amount"] else None
        sheet[f"G{row}"] = record["party"]
        sheet[f"H{row}"] = record["payment_reference"]
        sheet[f"I{row}"] = record["flags"]
        # The rows shipped with the workbook came from the Paper request, so
        # they carry its name and company code. Anything imported later
        # carries its own.
        sheet[f"Q{row}"] = "Paper Samples"
        sheet[f"R{row}"] = "GBKM"

        for column in ("C", "D", "E"):
            sheet[f"{column}{row}"].number_format = DATE_FMT
        sheet[f"F{row}"].number_format = AMOUNT_FMT

        for column, _, _ in SAMPLE_HEADERS:
            cell = sheet[f"{column}{row}"]
            cell.font = BODY_FONT
            cell.border = BOX
            if column in MACRO_COLUMNS:
                cell.fill = MACRO_FILL
            if column == "I":
                cell.font = Font(name=FONT, size=8, color="C00000")
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet[f"N{row}"].alignment = Alignment(horizontal="center")

        # Operator-controlled: blank or Yes runs the row, No skips it.
        include = sheet[f"P{row}"]
        include.value = "Yes"
        include.fill = INPUT_FILL
        include.font = INPUT_FONT
        include.alignment = Alignment(horizontal="center")

    last_row = header_row + len(rows)
    sheet.auto_filter.ref = f"A{header_row}:P{last_row}"

    include_validation = DataValidation(
        type="list", formula1='"Yes,No"', allow_blank=True, showDropDown=False
    )
    sheet.add_data_validation(include_validation)
    include_validation.add(f"P{header_row + 1}:P{last_row}")
    sheet[f"P{header_row}"].comment = Comment(
        "Set to No to leave a row out of the run. Blank counts as Yes. Use the "
        "autofilter on this column, or modSelect.IncludeOnlyMonth, to pick a subset.",
        "build_control_workbook.py",
    )

    total_row = last_row + 1
    sheet[f"E{total_row}"] = "Total"
    sheet[f"E{total_row}"].font = Font(name=FONT, size=10, bold=True)
    sheet[f"F{total_row}"] = f"=SUM(F{header_row + 1}:F{last_row})"
    sheet[f"F{total_row}"].font = Font(name=FONT, size=10, bold=True)
    sheet[f"F{total_row}"].number_format = AMOUNT_FMT
    sheet[f"F{total_row}"].border = Border(top=Side(style="double", color="1F3864"))
    # Plain text, not a formula -- a leading '=' here would make Excel try to
    # evaluate '56 samples' and yield an error.
    sheet[f"G{total_row}"] = f"{last_row - header_row} samples"
    sheet[f"G{total_row}"].font = NOTE_FONT


# --------------------------------------------------------------------------- #
# Log
# --------------------------------------------------------------------------- #

LOG_HEADERS = [
    ("A", "Timestamp", 19),
    ("B", "Sample #", 9),
    ("C", "SAP SID", 9),
    ("D", "Client", 8),
    ("E", "SAP user", 13),
    ("F", "Transaction", 12),
    ("G", "Action", 30),
    ("H", "Detail", 62),
    ("I", "Result", 11),
    ("J", "File written", 58),
]


def build_log(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Log")
    sheet.sheet_properties.tabColor = "7F7F7F"
    set_widths(sheet, {col: width for col, _, width in LOG_HEADERS})

    sheet["A2"] = "Audit trail"
    sheet["A2"].font = Font(name=FONT, size=14, bold=True, color="1F3864")
    sheet["A3"] = (
        "Appended by the macro, one row per action, in DRY RUN as well as EXTRACT. "
        "Keep this sheet with the extract -- it is the record of who pulled what, "
        "from which system, and when."
    )
    sheet["A3"].font = NOTE_FONT
    sheet.merge_cells("A3:J3")

    header_row = 4
    for column, title, _ in LOG_HEADERS:
        sheet[f"{column}{header_row}"] = title
    style_header(sheet, header_row, len(LOG_HEADERS))


# --------------------------------------------------------------------------- #
# Data Issues
# --------------------------------------------------------------------------- #

FLAG_EXPLANATIONS = {
    "date_stored_as_text": "The payment date was typed as text, not a date. Normalised on "
                           "extract; the auditor's own file still holds text.",
    "ref_whitespace_cleaned": "Leading or trailing spaces removed from the payment reference.",
    "party_whitespace_cleaned": "Leading, trailing or doubled spaces removed from the party name.",
    "no_named_beneficiary": "The request shows '-' as the party. These are the lines the "
                            "auditor's 'Payment to Supplier?' question actually turns on, "
                            "since the bank statement alone does not name a payee.",
    "ref_holds_a_party_name_not_a_transaction_description":
        "The payment-reference cell contains a party name. Confirm the intended "
        "transaction description with the auditor before searching on it.",
    "ref_typo_corrected_from:ACH PYMTS - LCL BULK FNG":
        "Transaction description misspelt ('FNG' for 'FNDG'). Corrected on extract.",
    "ref_unexpected": "Payment reference is neither of the two expected transaction "
                      "descriptions and does not look like a party name.",
    "month_column_disagrees_with_payment_date":
        "The 'month of payment' column points at a different month than the payment "
        "date. A human must decide which one drives the FEBAN date range.",
    "amount_missing": "No amount in the request.",
    "date_unparseable": "The payment date could not be read.",
}


def build_data_issues(workbook: Workbook, rows: list[dict]) -> None:
    sheet = workbook.create_sheet("Data Issues")
    sheet.sheet_properties.tabColor = "C00000"
    set_widths(sheet, {"A": 3, "B": 52, "C": 9, "D": 30, "E": 74})

    sheet["B2"] = "Data-quality flags in the auditor's request"
    sheet["B2"].font = Font(name=FONT, size=14, bold=True, color="1F3864")
    sheet["B3"] = (
        "Raised while reading 'Paper Samples'. Nothing here blocks the extract, but the "
        "starred rows are worth raising with the auditor before you send anything back."
    )
    sheet["B3"].font = NOTE_FONT
    sheet.merge_cells("B3:E3")

    counts: dict[str, list[int]] = {}
    for record in rows:
        for flag in (f.strip() for f in record["flags"].split(";") if f.strip()):
            counts.setdefault(flag, []).append(int(record["idx"]))

    header_row = 5
    for offset, title in enumerate(["Flag", "Count", "Sample #", "What it means"]):
        sheet.cell(row=header_row, column=2 + offset, value=title)
    style_header(sheet, header_row, 5)

    row = header_row + 1
    for flag, indices in sorted(counts.items(), key=lambda item: -len(item[1])):
        sheet.cell(row=row, column=2, value=flag).font = BODY_FONT
        sheet.cell(row=row, column=3, value=len(indices)).font = BODY_FONT
        sheet.cell(row=row, column=3).alignment = Alignment(horizontal="center")

        shown = ", ".join(str(i) for i in indices[:12])
        if len(indices) > 12:
            shown += f", ... (+{len(indices) - 12} more)"
        sheet.cell(row=row, column=4, value=shown).font = Font(name=FONT, size=9)
        sheet.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical="top")

        explanation = FLAG_EXPLANATIONS.get(flag, "Not recognised by the extract script.")
        sheet.cell(row=row, column=5, value=explanation).font = NOTE_FONT
        sheet.cell(row=row, column=5).alignment = Alignment(wrap_text=True, vertical="top")

        for col in range(2, 6):
            sheet.cell(row=row, column=col).border = BOX
        sheet.row_dimensions[row].height = 34
        row += 1

    row += 1
    sheet.cell(row=row, column=2, value="Findings outside the sample rows").font = SECTION_FONT
    row += 1
    findings = [
        "The auditor's month tabs already hold 55 pasted bank-statement screenshots, one "
        "per sample -- so the bank side of the evidence is done and the SAP side is what "
        "is missing.",
        "'Oct 25' holds 8 sample rows but only 7 screenshots, so one October line has no "
        "bank evidence attached. Worth confirming with the auditor which line.",
        "One October screenshot is for account 12343649 / IBAN GB49CITI18500812343649 in "
        "EUR, where every other screenshot is account 12343657 / GB27CITI18500812343657 "
        "in GBP. Either a second account is in scope or a screenshot was mis-pasted.",
        "Column G of 'Paper Samples', 'Payment to Supplier?', is blank on all 56 rows. "
        "That is the column this extract exists to let you answer.",
    ]
    for text in findings:
        sheet.cell(row=row, column=2, value="*").font = Font(name=FONT, size=10, bold=True)
        sheet.cell(row=row, column=3, value=text).font = NOTE_FONT
        sheet.cell(row=row, column=3).alignment = Alignment(wrap_text=True, vertical="top")
        sheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
        sheet.row_dimensions[row].height = 42
        row += 1


def main() -> None:
    here = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("-i", "--samples", type=Path, default=here.parent / "samples.csv")
    parser.add_argument(
        "-o", "--out", type=Path, default=here.parent / "FEBAN_Audit_Control.xlsx"
    )
    args = parser.parse_args()

    with args.samples.open(encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle))
    if not rows:
        raise SystemExit(f"error: {args.samples} holds no rows")

    workbook = Workbook()
    workbook.remove(workbook.active)
    build_control(workbook)
    build_screen_map(workbook)
    build_samples(workbook, rows)
    build_log(workbook)
    build_data_issues(workbook, rows)
    workbook.active = 0
    workbook.save(args.out)
    print(f"wrote {args.out}  ({len(rows)} samples, {len(workbook.sheetnames)} sheets)")


if __name__ == "__main__":
    main()
