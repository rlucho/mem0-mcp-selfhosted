#!/usr/bin/env python3
"""Extract and normalise the audit sample list from the auditor's request workbook.

Input : the auditor's 'Samples_Paper_*.xlsx' (sheet 'Paper Samples', header on rows 3-4)
Output: samples.csv -- one row per sample, cleaned, with data-quality flags

The source sheet is hand-maintained, so it carries the usual hand-maintained
problems: dates typed as text, stray leading/trailing spaces, a party name typed
into the payment-reference column, and a typo in one transaction description.
Everything this script repairs is recorded in the 'flags' column rather than
silently fixed, because the auditor's copy is the authoritative document.

Usage:
    python3 extract_samples.py <source.xlsx> [-o samples.csv]
"""

from __future__ import annotations

import argparse
import calendar
import csv
import datetime as dt
import re
import sys
from pathlib import Path

import openpyxl

SHEET = "Paper Samples"
FIRST_DATA_ROW = 5
COL_MONTH, COL_REF, COL_DATE, COL_AMOUNT, COL_PARTY = "B", "C", "D", "E", "F"

# Transaction descriptions we expect in the payment-reference column. Anything
# else is flagged rather than corrected -- see flag 'ref_unexpected'.
KNOWN_REFS = {
    "ACH PYMTS - LCL BULK FNDG",
    "ISSUE CHAPS PAYMENT",
}

# Known typos -> canonical form. Kept explicit so the mapping is reviewable.
REF_TYPO_FIXES = {
    "ACH PYMTS - LCL BULK FNG": "ACH PYMTS - LCL BULK FNDG",
}

# A payment reference that is really a party name. Detected, not guessed at.
PARTY_NAME_HINTS = ("DS SMITH", "SANTANDER", "LIMITED", "LTD")

TEXT_DATE_RE = re.compile(r"^\s*(\d{1,2})/(\d{1,2})/(\d{4})\s*$")


def squeeze(value: object) -> str:
    """Trim and collapse internal runs of whitespace."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def parse_date(value: object) -> tuple[dt.date | None, bool]:
    """Return (date, was_text). Handles real dates and 'dd/mm/yyyy' strings."""
    if isinstance(value, dt.datetime):
        return value.date(), False
    if isinstance(value, dt.date):
        return value, False
    match = TEXT_DATE_RE.match(str(value or ""))
    if match:
        day, month, year = (int(g) for g in match.groups())
        return dt.date(year, month, day), True
    return None, False


def month_key(date: dt.date) -> str:
    """'Sep 25' -- matches the per-month tab names in the auditor's workbook."""
    return f"{calendar.month_abbr[date.month]} {date.strftime('%y')}"


def extract(path: Path) -> list[dict]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    if SHEET not in workbook.sheetnames:
        sys.exit(f"error: sheet {SHEET!r} not found in {path.name}")
    sheet = workbook[SHEET]

    rows: list[dict] = []
    for excel_row in range(FIRST_DATA_ROW, sheet.max_row + 1):
        raw_month = sheet[f"{COL_MONTH}{excel_row}"].value
        raw_date = sheet[f"{COL_DATE}{excel_row}"].value
        raw_amount = sheet[f"{COL_AMOUNT}{excel_row}"].value
        if raw_month is None and raw_date is None and raw_amount is None:
            continue

        flags: list[str] = []

        month_of_payment, _ = parse_date(raw_month)
        payment_date, date_was_text = parse_date(raw_date)
        if date_was_text:
            flags.append("date_stored_as_text")
        if payment_date is None:
            flags.append("date_unparseable")

        # The statement-date range the macro will feed to FEBAN is derived from
        # the payment date, and cross-checked against the 'month of payment'
        # column. A mismatch means the sample list itself is inconsistent and a
        # human has to decide which column wins.
        if payment_date and month_of_payment:
            if (payment_date.year, payment_date.month) != (
                month_of_payment.year,
                month_of_payment.month,
            ):
                flags.append("month_column_disagrees_with_payment_date")

        ref_raw = sheet[f"{COL_REF}{excel_row}"].value
        ref = squeeze(ref_raw)
        if ref_raw is not None and str(ref_raw) != ref:
            flags.append("ref_whitespace_cleaned")
        if ref in REF_TYPO_FIXES:
            flags.append(f"ref_typo_corrected_from:{ref}")
            ref = REF_TYPO_FIXES[ref]
        if ref.upper() not in {r.upper() for r in KNOWN_REFS}:
            if any(hint in ref.upper() for hint in PARTY_NAME_HINTS):
                flags.append("ref_holds_a_party_name_not_a_transaction_description")
            else:
                flags.append("ref_unexpected")

        party_raw = sheet[f"{COL_PARTY}{excel_row}"].value
        party = squeeze(party_raw)
        if party_raw is not None and str(party_raw) != party:
            flags.append("party_whitespace_cleaned")
        if party == "-" or not party:
            party = ""
            flags.append("no_named_beneficiary")

        amount = float(raw_amount) if raw_amount is not None else None
        if amount is None:
            flags.append("amount_missing")

        statement_from = payment_date.replace(day=1) if payment_date else None
        statement_to = (
            payment_date.replace(
                day=calendar.monthrange(payment_date.year, payment_date.month)[1]
            )
            if payment_date
            else None
        )

        rows.append(
            {
                "idx": len(rows) + 1,
                "source_row": excel_row,
                "month_tab": month_key(payment_date) if payment_date else "",
                "statement_date_from": statement_from.isoformat() if statement_from else "",
                "statement_date_to": statement_to.isoformat() if statement_to else "",
                "payment_date": payment_date.isoformat() if payment_date else "",
                "amount": f"{amount:.2f}" if amount is not None else "",
                "party": party,
                "payment_reference": ref,
                "flags": "; ".join(flags),
            }
        )
    return rows


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("source", type=Path, help="auditor's request workbook (.xlsx)")
    parser.add_argument(
        "-o",
        "--out",
        type=Path,
        default=Path(__file__).resolve().parent.parent / "samples.csv",
        help="output CSV (default: ../samples.csv)",
    )
    args = parser.parse_args()

    rows = extract(args.source)
    if not rows:
        sys.exit("error: no sample rows found")

    args.out.parent.mkdir(parents=True, exist_ok=True)
    with args.out.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)

    flagged = sum(1 for r in rows if r["flags"])
    unnamed = sum(1 for r in rows if "no_named_beneficiary" in r["flags"])
    months: dict[str, int] = {}
    for row in rows:
        months[row["month_tab"]] = months.get(row["month_tab"], 0) + 1

    print(f"wrote {args.out}  ({len(rows)} samples)")
    print(f"  rows carrying at least one data-quality flag : {flagged}")
    print(f"  rows with no named beneficiary in the request : {unnamed}")
    print("  samples per month tab:")
    for month, count in months.items():
        print(f"    {month:<7} {count}")


if __name__ == "__main__":
    main()
