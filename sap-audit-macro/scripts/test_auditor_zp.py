"""Mirror of modImport.AuditorZpBelow, run against the real request files.

A follow-up request pastes an SAP extract under each sample naming the ZP the
auditor treated as the largest payment of that batch. The macro reads it so the
run can be checked against their answer instead of merely trusted.

The parse is by caption, never by position: the columns are not in the same
order on every block. In April'26 sample 3 'Clearing Document' is column M and
'Document Header Text' is N; two samples later they are the other way round. A
positional read would have returned the clearing document as though it were the
payment, and it would have looked entirely plausible.

Run from the repo root.
"""

import sys
from pathlib import Path

import openpyxl

UPLOADS = Path("/root/.claude/uploads/64ea30ba-b307-5571-8611-751275c5859a")

FILES = {
    "Follow-Up April to June 2026": "762024e1-FollowUp_Queries__SURL_Samples__April_to_June_2026260729_102054.265_2.xlsx",
    "SURL Samples - January-26": "48125468-SURL_Samples__January26260730_073243.407.xlsx",
    "SURL Samples - February-26": "a8321353-SURL_Samples__February26260730_073000.713.xlsx",
    "SURL Samples - March-26": "13de7826-SURL_Samples__March26260730_073430.512_1.xlsx",
}

# What the blocks in the follow-up file actually say, read by hand off the
# sheet. Anything the parser returns that is not in here is a bug in the
# parser, and anything in here it misses is a bug too.
EXPECTED = {
    "April Sample 3": "1637207",
    "April Sample 4": "1637661",
    "April Sample 5": "1638003",
    # Its block is the last thing on the sheet, and its columns are in a
    # different order from samples 3 and 4 -- M is 'Document Header Text' here
    # and N is 'Clearing Document', the other way round from four rows up.
    # Read positionally it returns 902499, the clearing document, which looks
    # entirely plausible and is wrong.
    "April Sample 6": "1638236",
    "May Sample 2": "1639297",
    "May Sample 3": "1639401",
    "May Sample 5": "1639749",
    "June Sample 2": "1640831",
    "June Sample 4": "1641172",
}


def cell(sheet, row, col):
    value = sheet.cell(row, col).value
    return "" if value is None else str(value).strip()


def sample_label(sheet, row, date_col):
    """SampleLabel: the '<Month> Sample <n>' text left of the dated columns."""
    for col in range(1, min(date_col - 1, 6) + 1):
        text = cell(sheet, row, col)
        if 6 <= len(text) <= 40 and "sample" in text.lower():
            return text
    return ""


def last_row_saying(sheet, label):
    """LastRowSaying: the block heading, which is always below the table."""
    for row in range(sheet.max_row, 0, -1):
        for col in range(1, 9):
            if cell(sheet, row, col).lower() == label.lower():
                return row
    return 0


def zp_in_block(sheet, block_row, wanted="ZP"):
    """ZpInBlock: find the header by caption, then the first row of that type."""
    doc_col = type_col = header_row = 0
    for row in range(block_row, min(block_row + 6, sheet.max_row) + 1):
        doc_col = type_col = 0
        for col in range(1, 21):
            caption = cell(sheet, row, col).lower()
            if "document number" in caption:
                doc_col = col
            if "document type" in caption:
                type_col = col
        if doc_col and type_col:
            header_row = row
            break
    if not header_row:
        return ""

    for row in range(header_row + 1, min(header_row + 30, sheet.max_row) + 1):
        if cell(sheet, row, type_col).upper() == wanted:
            return "".join(c for c in cell(sheet, row, doc_col) if c.isdigit())
        if "sample" in cell(sheet, row, 2).lower():
            return ""
    return ""


def find_date_col(sheet):
    """Good enough stand-in for FindColumns: the column headed 'Date'."""
    for row in range(1, min(20, sheet.max_row) + 1):
        for col in range(1, 15):
            if cell(sheet, row, col).lower() == "date":
                return col, row
    return 0, 0


def main():
    found, failures = {}, []

    for label, filename in FILES.items():
        path = UPLOADS / filename
        if not path.exists():
            print(f"  SKIP {label}: not on this machine")
            continue

        book = openpyxl.load_workbook(path, data_only=True)
        print(f"\n{label}")

        for sheet in book.worksheets:
            if sheet.title.startswith("DS_INTERNAL"):
                continue
            date_col, header_row = find_date_col(sheet)
            if not date_col:
                continue

            hits = 0
            for row in range(header_row + 1, sheet.max_row + 1):
                name = sample_label(sheet, row, date_col)
                if not name:
                    continue
                block_row = last_row_saying(sheet, name)
                if block_row <= row:
                    continue
                zp = zp_in_block(sheet, block_row)
                if zp:
                    found[name] = zp
                    hits += 1
                    print(f"    {name:<18} -> ZP {zp}")
            if hits == 0:
                print(f"    (sheet '{sheet.title}': no pasted extracts, as expected)")

    print("\n--- against the blocks read by hand ---")
    for name, want in EXPECTED.items():
        got = found.get(name)
        if got == want:
            print(f"  ok    {name:<18} {want}")
        else:
            print(f"  FAIL  {name:<18} expected {want}, parser said {got!r}")
            failures.append(name)

    extra = set(found) - set(EXPECTED)
    for name in sorted(extra):
        print(f"  FAIL  {name:<18} parser invented a block: {found[name]}")
        failures.append(name)

    print()
    if failures:
        print(f"{len(failures)} problem(s)")
        return 1
    print(f"all {len(EXPECTED)} blocks parsed correctly, none invented")
    return 0


if __name__ == "__main__":
    sys.exit(main())
